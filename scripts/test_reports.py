#!/usr/bin/env python3
"""Generate all report types using mock_data_final.xlsx.

Usage:
    cd payment-tracking-github-version
    python3 scripts/test_reports.py
"""

import os
import sys
from collections import defaultdict

import openpyxl

sys.path.insert(0, os.path.dirname(__file__))

from client_report  import generate_client_report
from weekly_report  import generate_weekly_report
from monthly_report import generate_monthly_report
from yearly_report  import generate_yearly_report, generate_general_report

XLSX    = os.path.join(os.path.dirname(__file__), '..', 'data', 'mock_data_final.xlsx')
OUT     = os.path.join(os.path.dirname(__file__), '..', 'reports', 'test')
COMPANY = 'Sample Construction Co.'


# ── Load data ─────────────────────────────────────────────────────────────────
def load_data():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws_c = wb['Clients']
    ws_r = wb['Receivables']

    c_h = [ws_c.cell(1, c).value for c in range(1, ws_c.max_column + 1)]
    r_h = [ws_r.cell(1, c).value for c in range(1, ws_r.max_column + 1)]

    clients = [
        {c_h[i]: ws_c.cell(r, i + 1).value for i in range(len(c_h))}
        for r in range(2, ws_c.max_row + 1)
    ]

    receivables = []
    for r in range(2, ws_r.max_row + 1):
        rec = {r_h[i]: ws_r.cell(r, i + 1).value for i in range(len(r_h))}
        # status and remaining_amount come from Excel formulas — calculate manually
        try:
            paid  = float(str(rec.get('paid_amount',  0) or 0).replace(',', '.'))
            total = float(str(rec.get('total_amount', 0) or 0).replace(',', '.'))
            rec['status']           = 'Paid' if paid >= total else ('Unpaid' if paid <= 0 else 'Partial')
            rec['remaining_amount'] = max(0.0, total - paid)
        except Exception:
            rec['status']           = 'Unpaid'
            rec['remaining_amount'] = float(str(rec.get('total_amount', 0) or 0).replace(',', '.'))
        receivables.append(rec)

    return clients, receivables


# ── Main test ─────────────────────────────────────────────────────────────────
def main():
    os.makedirs(OUT, exist_ok=True)
    clients, receivables = load_data()

    client_map       = {c['client_id']: c for c in clients}
    client_receivables = defaultdict(list)
    for r in receivables:
        client_receivables[r['client_id']].append(r)

    # Data summary
    n_p = sum(1 for r in receivables if r['status'] == 'Paid')
    n_t = sum(1 for r in receivables if r['status'] == 'Partial')
    n_u = sum(1 for r in receivables if r['status'] == 'Unpaid')
    print(f"Data: {len(clients)} clients, {len(receivables)} receivables "
          f"(Paid={n_p}, Partial={n_t}, Unpaid={n_u})\n")

    # ── Client reports ────────────────────────────────────────────────────────
    # Test scenarios: various payment distributions
    client_tests = [
        # (client_id, output_name, description)
        ('K004', 'test_client_partial.pdf',    'K004 — 2 Partial payments'),
        ('K038', 'test_client_mixed.pdf',      'K038 — 3 receivables, mixed (P+T+U)'),
        ('K003', 'test_client_perfect.pdf',    'K003 — 2 Paid (Perfect profile)'),
        ('K005', 'test_client_single.pdf',     'K005 — 1 receivable (Single)'),
    ]

    print("── Client Reports ──────────────────────────────────────────────")
    for cid, fname, description in client_tests:
        c   = client_map[cid]
        rs  = client_receivables[cid]
        out = os.path.join(OUT, fname)
        path = generate_client_report(c, rs, company_name=COMPANY, output_file=out)
        size_kb = os.path.getsize(path) // 1024
        print(f"  ✓ {fname:<35} {size_kb:>4} KB  ({description})")

    # ── Weekly reports ────────────────────────────────────────────────────────
    print("\n── Weekly Reports ──────────────────────────────────────────────")

    out_w = os.path.join(OUT, 'test_weekly.pdf')
    path = generate_weekly_report(clients, receivables, company_name=COMPANY, output_file=out_w)
    size_kb = os.path.getsize(path) // 1024
    print(f"  ✓ test_weekly.pdf                      {size_kb:>4} KB  (all clients, full data)")

    # Minimal subset (first 8 clients)
    mini_ids  = {c['client_id'] for c in clients[:8]}
    cli_mini  = [c for c in clients    if c['client_id'] in mini_ids]
    rec_mini  = [r for r in receivables if r['client_id'] in mini_ids]
    out_wm = os.path.join(OUT, 'test_weekly_minimal.pdf')
    path = generate_weekly_report(cli_mini, rec_mini, company_name=COMPANY, output_file=out_wm)
    size_kb = os.path.getsize(path) // 1024
    print(f"  ✓ test_weekly_minimal.pdf              {size_kb:>4} KB  (8 clients, small dataset)")

    # ── Monthly report ────────────────────────────────────────────────────────
    print("\n── Monthly Report ──────────────────────────────────────────────")

    out_m = os.path.join(OUT, 'test_monthly.pdf')
    path = generate_monthly_report(clients, receivables, company_name=COMPANY, output_file=out_m)
    size_kb = os.path.getsize(path) // 1024
    print(f"  ✓ test_monthly.pdf                     {size_kb:>4} KB  (all clients, January 2026)")

    # ── Yearly reports ────────────────────────────────────────────────────────
    print("\n── Yearly Reports ──────────────────────────────────────────────")

    out_y26 = os.path.join(OUT, 'test_yearly_2026.pdf')
    path = generate_yearly_report(clients, receivables, year=2026,
                                  company_name=COMPANY, output_file=out_y26)
    size_kb = os.path.getsize(path) // 1024
    print(f"  ✓ test_yearly_2026.pdf                 {size_kb:>4} KB  (all clients, 2026)")

    out_y25 = os.path.join(OUT, 'test_yearly_2025.pdf')
    path = generate_yearly_report(clients, receivables, year=2025,
                                  company_name=COMPANY, output_file=out_y25)
    size_kb = os.path.getsize(path) // 1024
    print(f"  ✓ test_yearly_2025.pdf                 {size_kb:>4} KB  (all clients, 2025)")

    # ── General report (all periods) ──────────────────────────────────────────
    print("\n── General Report (All Periods) ────────────────────────────────")

    out_g = os.path.join(OUT, 'test_general_report.pdf')
    path = generate_general_report(clients, receivables, company_name=COMPANY, output_file=out_g)
    size_kb = os.path.getsize(path) // 1024
    print(f"  ✓ test_general_report.pdf              {size_kb:>4} KB  (all clients, all periods)")

    print(f"\nAll PDFs → {OUT}/")


if __name__ == '__main__':
    main()
